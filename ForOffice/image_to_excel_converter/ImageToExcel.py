import openpyxl
from PIL import Image
from openpyxl.styles import PatternFill
import numpy as np
import os


class ImageToExcel:
    def __init__(self, file, save='sample.xlsx', indent_top=0, indent_left=0):
        self.file = file
        self.save = save
        self.indent_top = indent_top
        self.indent_left = indent_left

    def image_to_rgba_array(self):
        img = Image.open(self.file).convert('RGBA')
        arr = np.array(img)
        return arr

    def arr_to_hex_argb_color_matr(self, arr):
        fstring = '{0:02x}{1:02x}{2:02x}'
        _, file_ext = os.path.splitext(self.file)
        if file_ext == '.png':
            fstring = '{3:02x}' + fstring

        color_matr = [list(map(lambda x: fstring.format(*x), row))
                      for row in arr]
        return color_matr

    def create_table_with_img(self, clr_matr):
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(len(clr_matr)):
            for j, clr in enumerate(clr_matr[i]):
                if len(clr) == 8 and clr.startswith('00'):
                    continue
                clrFill = PatternFill(start_color=clr,
                                      end_color=clr,
                                      fill_type='solid')
                ws.cell(row=i + 1 + self.indent_top, column=j +
                        1 + self.indent_left).fill = clrFill
        wb.save(self.save)

    def img_to_excel(self):
        arr = self.image_to_rgba_array()
        color_matr = self.arr_to_hex_argb_color_matr(arr)
        self.create_table_with_img(color_matr)
