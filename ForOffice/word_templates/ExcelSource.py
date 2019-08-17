import openpyxl


class ExcelSource:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(filename=self.file_path)

    def get_data(self, sheet_name=None):
        if sheet_name is not None:
            sheet = self.wb[sheet_name]
        else:
            sheet = self.wb[self.get_sheets()[0]]
        data = []
        for i in range(1, self._count(sheet, True) - 1):
            data.append({
                sheet.cell(row=1, column=j).value:
                sheet.cell(row=i + 1, column=j).value
                for j in range(1, self._count(sheet, False))
            })
        return data

    @staticmethod
    def _count(sheet, flag):
        # True - rows
        # False - columns
        n, m = 1, 1
        cell = sheet.cell(row=n, column=m).value
        while cell:
            if flag:
                n += 1
            else:
                m += 1
            cell = sheet.cell(row=n, column=m).value
        return n if flag else m

    def get_sheets(self):
        return self.wb.sheetnames

    def get_fields(self, sheet_name):
        sheet = self.wb[sheet_name]
        return [sheet.cell(row=1, column=j).value
                for j in range(1, self._count(sheet, False))]
